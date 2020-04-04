import openpyxl

def get_num_of_rows(path):
    workbook = openpyxl.load_workbook(path + "\insta_search.xlsx")
    Sheet = workbook.worksheets[0]
    return Sheet.max_row

def read_excel(path, start_row, end_row):
    workbook = openpyxl.load_workbook(path + "\insta_search.xlsx")
    Sheet = workbook.worksheets[0]

    id_list = []

    for i in range(start_row, end_row + 1):
        id_list.append( Sheet.cell(row=i, column=1).value )

    return id_list

def write_excel(path, row_num, col_num, value):
    try:
        workbook = openpyxl.load_workbook(path + "\insta_search.xlsx")
        sheet = workbook.worksheets[0]

        for eachvalue in value:
            sheet.cell(row=row_num, column=col_num).value = eachvalue
            row_num = row_num + 1

        workbook.save(path + "\insta_search.xlsx")
    except PermissionError:
        print("Error Occured : Close the Excel Window before executing code")